import openpyxl
from openpyxl import *
import pandas as pd
import os

paths = ['2022_data', '2021_data', '2020_data', '2019_data']

combined_data = []
for path in paths:
    workbook = openpyxl.load_workbook(f'C:/Users/User/Desktop/Kens_Things/Elula_Project/{path}.xlsx', read_only=True, data_only=True)
    print("workbook has been opened")

    # contains all sheetnames)
    sheetnames = workbook.sheetnames
    workbook.close()

    dict = {}
    xlsx_path = f'C:/Users/User/Desktop/Kens_Things/Elula_Project/{path}.xlsx'
    xlsx_file = pd.ExcelFile(f'C:/Users/User/Desktop/Kens_Things/Elula_Project/{path}.xlsx')
    for i in range(0, len(sheetnames)):
        key = f'df{i + 1}'
        value = pd.read_excel(xlsx_file, sheetnames[i])
        dict[key] = value

    for key, value in dict.items():
        # this is setting the new headers to the [0] row index.
        dict[key].columns = dict[key].iloc[0]
        # this print only those from index 1 onwards. Setting drop=True means reset_Index will delete the index
        df = dict[key].iloc[1:].reset_index(drop=True)
        # gets rid of the indexed rows on axis=0
        df = df.drop(index=[0, 1, 7, 8, 19, 21, 30, 31, 33, 36, 43, 48, 49, 50, 54, 55], axis=0)
        # this gets rid of the columns named 'H' and 'Mortgage interest rates' on axis=1. Resets the index from 0 onwards
        df2 = df.drop(['H', 'Mortgage interest rates'], axis=1).reset_index(drop=True)
        df2.to_csv(f'data{key}.csv', index=False)

    yearly_data = []
    for key, value in dict.items():
        df = pd.read_csv(f'data{key}.csv')
        # out of all of the columns, gets rid of the one that contains '^Unnamed: 8'
        df = df.loc[:, ~df.columns.str.contains('^Unnamed: 8')]
        # renames all the columns
        df.rename(columns={'Unnamed: 0': '1', 'Unnamed: 1': '2', 'Unnamed: 2': '3', 'Unnamed: 3': '4', 'Unnamed: 4': '5',
                           'Unnamed: 5': '6', 'Unnamed: 6': '7', 'Unnamed: 7': '8'}, inplace=True)
        # Turns all of the dataframe values into one long list of arrays
        df_list = df.values.tolist()
        # so for each array [] in this list, print every element so that our final list is one long list of only numbers
        for list in df_list:
            for x in range(len(list)):
                yearly_data.append(list[x])
        os.remove(f'data{key}.csv')

    for x in range(len(yearly_data)):
        combined_data.append(yearly_data[x])

# create a new data frame that has only one column called 'Rate' with all of the rates as the data under it
combined_dataframe = pd.DataFrame(combined_data, columns=['Rate'])
print(combined_dataframe)
combined_dataframe.to_csv(f'Final_data.csv')