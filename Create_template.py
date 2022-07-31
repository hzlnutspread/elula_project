import openpyxl
from openpyxl import *
import pandas as pd
from datetime import datetime, timedelta

paths = ['2022_data', '2021_data', '2020_data', '2019_data']
for path in paths:
    workbook = openpyxl.load_workbook(f'C:/Users/User/Desktop/Kens_Things/Elula_Project/{path}.xlsx', read_only=True, data_only=True)
    print("workbook has been opened")
    # contains all sheetnames)
    sheetnames = workbook.sheetnames
    workbook.close()

    final_path = f'C:/Users/User/Desktop/Kens_Things/Elula_Project/Final_frame.csv'
    for date in sheetnames:
        df = pd.read_csv('C:/Users/User/Desktop/Kens_Things/Elula_Project/Test_data.csv')
        df['Date'] = df['Date'].replace(['15-Oct-21'], f'{date}')
        df = df.drop('Rate', axis=1)
        print(df)
        df.to_csv(final_path, mode='a', index=False, header=False)